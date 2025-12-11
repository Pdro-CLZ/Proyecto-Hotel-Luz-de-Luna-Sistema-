from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from reservas.models import Reserva, FechaReservada
import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from io import BytesIO
from xhtml2pdf import pisa

from administracion.decorators import rol_requerido


ESTADOS_HABITACION = [
    ('ocupada', 'Ocupada'),
    ('disponible', 'Disponible'),
]


@rol_requerido("Administrador", "Empleado_Nivel1")
def reporte_reservas(request):

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado_habitacion = request.GET.get('estado_habitacion')  # (por si luego lo usás)
    canal_reservacion = request.GET.get('canal_reservacion')
    exportar = request.GET.get('exportar')

    reservas = Reserva.objects.select_related('habitacion', 'cliente').all()

    fi = ff = None

    # Filtro por rango de fechas
    if fecha_inicio and fecha_fin:
        try:
            fi = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            ff = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d').date()

            if fi <= ff:
                reservas = reservas.filter(
                    fecha_inicio__gte=fi,
                    fecha_fin__lte=ff
                )
        except ValueError:
            # Si la fecha viene mala, simplemente no se filtra
            pass

    # Filtro por canal de reservación
    if canal_reservacion:
        reservas = reservas.filter(canal_reservacion=canal_reservacion)

    total_reservas = reservas.count()
    total_monto = reservas.aggregate(total=Sum('total'))['total'] or 0

    # Texto legible del canal para usar en el PDF
    canal_texto = "Todos"
    if canal_reservacion:
        canal_dict = dict(Reserva.CANALES_RESERVACION)
        canal_texto = canal_dict.get(canal_reservacion, canal_reservacion)

    # ===== Exportar =====
    if exportar == 'excel':
        return exportar_reservas_excel(reservas)

    if exportar == 'pdf':
        return exportar_reservas_pdf(
            reservas_qs=reservas,
            total_reservas=total_reservas,
            total_monto=total_monto,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            canal_texto=canal_texto,
        )

    # Render normal HTML
    context = {
        'reservas': reservas,
        'total_reservas': total_reservas,
        'total_monto': total_monto,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estado_habitacion': estado_habitacion,
        'canal_reservacion': canal_reservacion,
        'opciones_estado': ESTADOS_HABITACION,
        'opciones_canal': Reserva.CANALES_RESERVACION,
    }
    return render(request, 'reporteria/reporte_reservas.html', context)


def exportar_reservas_excel(reservas_qs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas"

    columnas = [
        'ID', 'Cliente', 'Habitación', 'Fecha inicio',
        'Fecha fin', 'Total', 'Canal', 'Método de pago'
    ]
    ws.append(columnas)

    for r in reservas_qs:
        ws.append([
            r.id,
            str(r.cliente),
            r.habitacion.nombre,
            r.fecha_inicio.strftime('%Y-%m-%d'),
            r.fecha_fin.strftime('%Y-%m-%d'),
            float(r.total),
            r.get_canal_reservacion_display(),
            r.get_metodo_pago_display(),
        ])

    for col_num, _ in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_reservas.xlsx'
    wb.save(response)
    return response


def exportar_reservas_pdf(reservas_qs, total_reservas, total_monto,
                          fecha_inicio, fecha_fin, canal_texto):
    # Armar filas de la tabla
    filas_html = ""
    for r in reservas_qs:
        filas_html += f"""
        <tr>
            <td>{r.id}</td>
            <td>{r.cliente}</td>
            <td>{r.habitacion.nombre}</td>
            <td>{r.fecha_inicio}</td>
            <td>{r.fecha_fin}</td>
            <td style="text-align:right;">${float(r.total):,.2f}</td>
            <td>{r.get_canal_reservacion_display()}</td>
            <td>{r.get_metodo_pago_display()}</td>
        </tr>
        """

    if not filas_html:
        filas_html = """
        <tr>
            <td colspan="8">Sin registros disponibles para los filtros seleccionados.</td>
        </tr>
        """

    # Texto de rango de fechas
    if fecha_inicio and fecha_fin:
        rango_texto = f"{fecha_inicio} al {fecha_fin}"
    else:
        rango_texto = "Todas"

    # HTML básico del reporte
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
    <meta charset="UTF-8">
    <title>Reporte de Reservas</title>
    <style>
      body {{ font-family: DejaVu Sans, sans-serif; font-size: 11px; }}
      h2 {{ text-align: center; margin-bottom: 10px; }}
      p {{ margin: 2px 0; }}
      table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
      th, td {{ border: 1px solid #444; padding: 4px; text-align: center; }}
      th {{ background-color: #eee; }}
      .right {{ text-align: right; }}
    </style>
    </head>
    <body>
      <h2>Reporte de Reservas</h2>

      <p><strong>Rango de fechas:</strong> {rango_texto}</p>
      <p><strong>Canal de reservación:</strong> {canal_texto}</p>
      <p><strong>Total de reservas:</strong> {total_reservas} &nbsp;&nbsp;
         <strong>Monto total:</strong> ${float(total_monto):,.2f}</p>

      <table>
        <thead>
          <tr>
            <th>ID</th>
            <th>Cliente</th>
            <th>Habitación</th>
            <th>Fecha inicio</th>
            <th>Fecha fin</th>
            <th>Total</th>
            <th>Canal</th>
            <th>Método de pago</th>
          </tr>
        </thead>
        <tbody>
          {filas_html}
        </tbody>
      </table>
    </body>
    </html>
    """

    result = BytesIO()
    pisa.CreatePDF(html, dest=result)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_reservas.pdf"'
    return response
